"""動作検証スクリプト（pytest 非依存）。

    python tests/verify.py [path/to/workbook.xlsx]

引数で xlsx を渡すとオフライン検証。省略時は既定シートをライブ取得して検証する。
ユニット検証（純関数）＋統合検証（実データのロード・絞り込み・整形）を実行。
"""

from __future__ import annotations

import sys
from pathlib import Path

try:
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:  # noqa: BLE001
    pass

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from publication_summarizer import (  # noqa: E402
    AuthorMatcher,
    DEFAULT_SHEET_ID,
    RECORD_TYPES,
    by_authors,
    by_fiscal_year,
    load_publications,
    load_templates,
    parse_roster,
    render_records,
)
from publication_summarizer.formatter import (  # noqa: E402
    AuthorStyle,
    clean_number,
    render_authors,
    render_one,
    strip_title_wrap,
    _cleanup,
)
from publication_summarizer.loader import (  # noqa: E402
    _fiscal_year,
    _has_cjk,
    _parse_date,
    load_roster_sheet,
    load_workbook_bytes,
)
from publication_summarizer.filters import by_peer_reviewed, by_scope, by_invited  # noqa: E402
from publication_summarizer.i18n import rt_label, tr  # noqa: E402
from publication_summarizer.roster import Member, split_authors  # noqa: E402
from publication_summarizer.schema import BILINGUAL_FIELDS, display_fields  # noqa: E402
from tests.test_form_fields import form_field_tests, template_header_tests  # noqa: E402

PASS, FAIL = 0, 0


def check(name: str, cond: bool, detail: str = "") -> None:
    global PASS, FAIL
    if cond:
        PASS += 1
        print(f"  [OK]   {name}")
    else:
        FAIL += 1
        print(f"  [FAIL] {name}  {detail}")


def unit_tests() -> None:
    print("[unit] clean_number")
    check("30.0 -> 30", clean_number("30.0") == "30")
    check("871.0 -> 871", clean_number(871.0) == "871")
    check("range kept", clean_number("1002-1010") == "1002-1010")
    check("empty", clean_number(None) == "")

    print("[unit] fiscal year")
    check("Apr -> same FY", _fiscal_year(_parse_date("2025/04/10")) == 2025)
    check("Jan -> prev FY", _fiscal_year(_parse_date("2026/01/16")) == 2025)
    check("slash short", _fiscal_year(_parse_date("2025/4/21")) == 2025)

    print("[unit] cleanup of empty fields")
    check("empty parens removed", "()" not in _cleanup("Title (). 2020"))
    check("dangling sep collapsed", _cleanup("a. b. 2020;():.") == "a. b. 2020")

    print("[unit] i18n")
    check("ja/en differ", tr("app_title", "ja") != tr("app_title", "en"))
    check("rt_label en", rt_label("paper", "en") == "Original Papers / English Reviews")

    print("[unit] author matching (synthetic)")
    members = [
        Member(ja="林 直子", last="Hayashi", first="Naoko"),
        Member(ja="宮崎 慎一", last="Miyazaki", first="Shinichi"),
    ]
    m = AuthorMatcher(members, threshold=85)
    hayashi = members[0]
    check("full name", m.matches_member("Naoko Hayashi", hayashi))
    check("pubmed Lastname I", m.matches_member("Hayashi N", hayashi))
    check("with period", m.matches_member("Naoko Hayashi.", hayashi))
    check("japanese", m.matches_member("林 直子", hayashi))
    check("not other person", not m.matches_member("Taro Yamada", hayashi))
    # 精度: 同姓でも名（頭文字）が違えば別人として不一致にする。
    check("同姓別人(Yu Hayashi)は不一致", not m.matches_member("Yu Hayashi", hayashi))
    check("同姓別人(Kenji Hayashi)は不一致", not m.matches_member("Kenji Hayashi", hayashi))
    check("姓のみ(特定不能)は不一致", not m.matches_member("Hayashi", hayashi))
    check(
        "record has member (Miyazaki S を含む)",
        m.record_has_any("Okamura H, Yasugaki S, Miyazaki S", members),
    )
    check(
        "record has none (同姓別人のみ)",
        not m.record_has_any("Okamura H, Yasugaki S, Hayashi Y", members),
    )

    print("[unit] resolve_member（イニシャル整形用の名簿照合）")
    check("full→member", m.resolve_member("Naoko Hayashi") is hayashi)
    check("japanese→member", m.resolve_member("林 直子") is hayashi)
    check("外部共著者は None", m.resolve_member("Taro Yamada") is None)

    print("[unit] split_authors（区切り・et al. 除去）")
    check("カンマ区切り", split_authors("A, B, C") == ["A", "B", "C"])
    check("and / & / ; も区切る",
          split_authors("A, B and C; D & E") == ["A", "B", "C", "D", "E"])
    check("末尾 et al. を除去", split_authors("Smith J, et al.") == ["Smith J"])
    check("末尾 ほかN名 を除去", split_authors("宮崎 慎一, ほか3名") == ["宮崎 慎一"])
    check("and を含む姓を割らない", split_authors("Anderson J, Bond K") == ["Anderson J", "Bond K"])
    check("発表者の丸印を除去", split_authors("○山田 太郎, 林 直子") == ["山田 太郎", "林 直子"])
    check("丸印＋空白も除去", split_authors("◯ 山田 太郎, 〇林 直子") == ["山田 太郎", "林 直子"])


def author_style_tests() -> None:
    """著者整形（人数省略・自己強調・言語別省略語・二重太字回避）の検証。"""
    print("[unit] render_authors（省略・言語別省略語）")
    five = "A, B, C, D, E"
    check("全員(max=0)", render_authors(five, AuthorStyle()) == "A, B, C, D, E")
    check("上位3+et al(en)",
          render_authors(five, AuthorStyle(max_authors=3), "en") == "A, B, C, et al.")
    check("上位3+ほか(ja)",
          render_authors(five, AuthorStyle(max_authors=3), "ja") == "A, B, C, ほか")
    check("ほかN名(ja, count)",
          render_authors(five, AuthorStyle(max_authors=3, etal_count=True), "ja") == "A, B, C, ほか2名")
    check("省略語の明示指定",
          render_authors(five, AuthorStyle(max_authors=2, etal="…ほか"), "ja") == "A, B, …ほか")

    print("[unit] render_authors（自己強調 × 省略の協調）")
    hl = lambda t: t == "Miyazaki S"  # noqa: E731
    inr = "Miyazaki S, A, B, C, D"
    check("強調が範囲内(plain)",
          render_authors(inr, AuthorStyle(max_authors=3), "en", hl) == "Miyazaki S, A, B, et al.")
    check("強調が範囲内(md太字)",
          render_authors(inr, AuthorStyle(max_authors=3), "en", hl, markdown=True)
          == "**Miyazaki S**, A, B, et al.")
    mid = "A, B, C, D, E, Miyazaki S, Z"
    check("強調が中間→…+et al",
          render_authors(mid, AuthorStyle(max_authors=3), "en", hl) == "A, B, C, …, Miyazaki S, et al.")
    tail = "A, B, C, D, E, Miyazaki S"
    check("強調が末尾→…のみ(et alなし)",
          render_authors(tail, AuthorStyle(max_authors=3), "en", hl) == "A, B, C, …, Miyazaki S")
    check("emphasis=none は太字にしない",
          render_authors(inr, AuthorStyle(max_authors=3, emphasis="none"), "en", hl, markdown=True)
          == "Miyazaki S, A, B, et al.")
    check("plain は強調マーカーなし", "**" not in render_authors(mid, AuthorStyle(max_authors=3), "en", hl))

    print("[unit] 二重太字回避（authors を bold 指定 + 自己強調 bold）")
    rec = {"type": "paper", "label": "x", "authors_raw": "Miyazaki S, A, B",
           "title_en": "T", "title_ja": "T", "date": None}
    out = render_one(rec, "{authors}. {title}", (), AuthorStyle(emphasis="bold"),
                     True, {"authors"}, set(), "en", hl)
    check("**** が出ない", "****" not in out, out)
    check("内側の強調は残る", "**Miyazaki S**" in out, out)

    print("[unit] render_authors（イニシャル整形）")
    resolver = lambda t: {"Naoko Hayashi": "Hayashi N", "林 直子": "Hayashi N"}.get(t)  # noqa: E731
    check("照合できた著者をイニシャル化",
          render_authors("Naoko Hayashi, Taro Yamada", AuthorStyle(initials=True), "en",
                         name_resolver=resolver) == "Hayashi N, Taro Yamada")
    check("initials=False なら変換しない",
          render_authors("Naoko Hayashi", AuthorStyle(initials=False), "en",
                         name_resolver=resolver) == "Naoko Hayashi")
    check("強調と併用（イニシャル化した名を太字）",
          render_authors("林 直子, A", AuthorStyle(initials=True), "ja",
                         (lambda t: t == "林 直子"), markdown=True, name_resolver=resolver)
          == "**Hayashi N**, A")

    print("[unit] strip_title_wrap（タイトルの括弧・引用符外し）")
    check("和括弧を外す", strip_title_wrap("「睡眠の科学」") == "睡眠の科学")
    check("英引用符を外す", strip_title_wrap('"Sleep science"') == "Sleep science")
    check("途中の括弧は保持", strip_title_wrap("TNF-α (review)") == "TNF-α (review)")
    check("入れ子も外す", strip_title_wrap("（「題」）") == "題")
    rec2 = {"type": "paper", "label": "x", "authors_raw": "Yamada T",
            "title_en": "「Wrapped」", "title_ja": "「Wrapped」", "date": None}
    out2 = render_one(rec2, "{title}", (), AuthorStyle(), False, set(), set(), "en")
    check("render_one でも括弧外し", out2 == "Wrapped", out2)


def v2_tests() -> None:
    """v2（二ヶ国語・査読正規化・upgrade）の純関数・往復検証。"""
    import pandas as pd
    from openpyxl import Workbook, load_workbook

    print("[v2] 言語判定 (_has_cjk)")
    check("CJK→True", _has_cjk("睡眠とストレス"))
    check("ASCII→False", not _has_cjk("Sleep and stress"))

    print("[v2] display_fields は base へ集約")
    df_paper = display_fields("paper")
    check("title が base で含まれる", "title" in df_paper and "title_ja" not in df_paper)
    check("journal_abbr が base", "journal_abbr" in df_paper)

    print("[v2] formatter の言語解決＋フォールバック")
    rec = {
        "type": "paper", "label": "x", "authors_raw": "Yamada T",
        "title_ja": "マウスのレム睡眠", "title_en": "REM sleep in mice",
        "journal_ja": "", "journal_en": "Sci Rep",
        "date": _parse_date("2026/04/10"),
    }
    pat = "{authors}. {title}. {journal}. {year}"
    en = render_one(rec, pat, (), AuthorStyle(), False, set(), set(), "en")
    ja = render_one(rec, pat, (), AuthorStyle(), False, set(), set(), "ja")
    check("en はタイトル英語", "REM sleep in mice" in en, en)
    check("ja はタイトル日本語", "マウスのレム睡眠" in ja, ja)
    check("journal はja空→en採用(フォールバック)", "Sci Rep" in ja and "Sci Rep" in en)

    print("[v2] 査読判定の正規化")
    pr = pd.DataFrame({"type": ["paper", "paper", "paper"],
                       "peer_reviewed": ["査読あり", "〇", "査読なし"]})
    kept = by_peer_reviewed(pr, True)
    check("査読あり/〇 を採用, 査読なし を除外", len(kept) == 2, f"n={len(kept)}")

    # 査読列を持たない種別（発表）は査読フィルタで消えない（素通し）。
    pr2 = pd.DataFrame({"type": ["paper", "paper", "presentation"],
                        "peer_reviewed": ["査読あり", "査読なし", None]})
    kept2 = by_peer_reviewed(pr2, True)
    check("発表は査読フィルタで素通し", set(kept2["type"]) == {"paper", "presentation"}
          and len(kept2) == 2, f"types={list(kept2['type'])}")

    print("[v2] 国内/国際フィルタ（by_scope）")
    sc = pd.DataFrame({"type": ["presentation", "presentation", "presentation", "paper"],
                       "scope": ["国内", "国際", "International", None]})
    intl = by_scope(sc, "国際")
    check("国際を選ぶと国際のみ（論文は素通し, 空scopeの発表は除外）",
          set(intl["type"]) == {"presentation", "paper"} and len(intl) == 3,
          f"n={len(intl)} types={list(intl['type'])}")
    check("すべて（空）は全件", len(by_scope(sc, "")) == 4)
    dom = by_scope(sc, "国内")
    check("国内を選ぶと国内のみ＋論文素通し", len(dom) == 2)

    print("[v2] 招待フィルタ（by_invited）")
    inv = pd.DataFrame({"type": ["presentation", "presentation", "presentation", "paper"],
                        "invited": ["招待あり", "招待なし", None, None]})
    yes = by_invited(inv, "招待あり")
    check("招待ありを選ぶと招待ありのみ（論文は素通し, 空invitedの発表は除外）",
          set(yes["type"]) == {"presentation", "paper"} and len(yes) == 2,
          f"n={len(yes)} types={list(yes['type'])}")
    check("すべて（空）は全件", len(by_invited(inv, "")) == 4)
    check("招待なしを選ぶと招待なしのみ＋論文素通し", len(by_invited(inv, "招待なし")) == 2)
    # invited 列を持たない種別（論文のみ）は招待フィルタで消えない（素通し）。
    inv2 = pd.DataFrame({"type": ["paper", "paper"], "invited": [None, None]})
    check("invited列があっても発表以外は素通し", len(by_invited(inv2, "招待あり")) == 2)

    print("[v2] ingest --from upgrade 往復")
    import importlib.util
    spec = importlib.util.spec_from_file_location(
        "ingest_to_canonical",
        Path(__file__).resolve().parent.parent / "scripts" / "ingest_to_canonical.py")
    ing = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(ing)

    print("[v2] 重複検出（DOI正規化・月違い・タイトル一致を横断）")
    check("JP日付パース 2021年12月", str(_parse_date("2021年12月").date()) == "2021-12-01")
    check("DOI接頭辞除去", ing._clean_doi("doi: 10.1/ABC") == "10.1/abc"
          and ing._clean_doi("https://doi.org/10.2/x") == "10.2/x")
    ddir = Path(__file__).resolve().parent
    dpath = ddir / "_v2_dup.xlsx"
    try:
        import pandas as pd2
        # 論文: DOI有無の混在・月違い(4月/7月)・末尾ピリオド差 でも同一タイトルなら集約。
        a = {"date": "2025/7", "title": "Circular RNA biomarker", "authors": "Miyazaki S"}
        b = {"date": "2025/07/15", "title": "Circular RNA biomarker", "authors": "Miyazaki S",
             "doi": "10.1/abc"}
        d = {"date": "2025/4/21", "title": "Circular RNA biomarker.", "authors": "Miyazaki S",
             "doi": "doi: 10.1/ABC"}
        ing.write_canonical(dpath, {**{rt: [] for rt in ing.CANONICAL_FIELDS}, "paper": [a]},
                            "paste", "未確認")
        ing.write_canonical(dpath, {**{rt: [] for rt in ing.CANONICAL_FIELDS}, "paper": [b, dict(a), d]},
                            "paste", "未確認", append_to=dpath)
        ddf = pd2.read_excel(dpath, sheet_name="Original Papers")
        check("論文: 同一タイトルは1件に集約（月違い・DOIゆれ横断）", len(ddf) == 1, f"rows={len(ddf)}")

        # 発表: 同名講演を別日に行うのは正当 → タイトル一致のみでは集約しない。
        p1 = {"date": "2024/8", "title": "睡眠の制御機構", "authors": "Miyazaki S"}
        p2 = {"date": "2023/8", "title": "睡眠の制御機構", "authors": "Miyazaki S"}
        ing.write_canonical(dpath, {**{rt: [] for rt in ing.CANONICAL_FIELDS}, "presentation": [p1, p2]},
                            "paste", "未確認", append_to=dpath)
        pdf = pd2.read_excel(dpath, sheet_name="presentations")
        check("発表: 同名でも別日なら2件保持", len(pdf) == 2, f"rows={len(pdf)}")
    finally:
        if dpath.exists():
            dpath.unlink()

    tmp_in = Path(__file__).resolve().parent / "_v2_old.xlsx"
    tmp_out = Path(__file__).resolve().parent / "_v2_new.xlsx"
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Original Papers"
        old_header = ["record_id", "status", "submitter", "source", "created_at",
                      "date", "category", "peer_reviewed", "authors", "title",
                      "journal", "journal_abbr", "volume", "issue", "pages", "doi"]
        ws.append(old_header)
        ws.append(["PAP-0001", "確認済", "山田", "legacy", "2026/01/01",
                   "2026/04/10", "原著論文", "査読あり", "Yamada T",
                   "REM sleep in mice", "Sci Rep", "Sci. Rep.", "16", "1", "871", "10.x/y"])
        ws.append(["PAP-0002", "確認済", "林", "legacy", "2026/01/01",
                   "2025/09/20", "原著論文", "査読あり", "Hayashi N",
                   "マウスの社会的順位", "実験医学", "実医", "10", "2", "33", ""])
        wb.save(tmp_in)

        counts = ing.upgrade_canonical(tmp_in, tmp_out)
        check("upgrade で2件移行", counts.get("paper") == 2, f"counts={counts}")
        rwb = load_workbook(tmp_out)
        rws = rwb["Original Papers"]
        hdr = [c.value for c in rws[1]]
        check("note 列が追加", "note" in hdr)
        check("title_ja/title_en 列が追加", "title_ja" in hdr and "title_en" in hdr)
        rows = list(rws.iter_rows(min_row=2, values_only=True))
        idx = {h: i for i, h in enumerate(hdr)}
        en_row = next(r for r in rows if r[idx["record_id"]] == "PAP-0001")
        ja_row = next(r for r in rows if r[idx["record_id"]] == "PAP-0002")
        check("英語タイトル→title_en", en_row[idx["title_en"]] == "REM sleep in mice"
              and not str(en_row[idx["title_ja"]] or "").strip())
        check("和タイトル→title_ja", ja_row[idx["title_ja"]] == "マウスの社会的順位"
              and not str(ja_row[idx["title_en"]] or "").strip())
        check("status/record_id 保持", en_row[idx["status"]] == "確認済")
    finally:
        for p in (tmp_in, tmp_out):
            if p.exists():
                p.unlink()


def paste_tests() -> None:
    """researchmap 等のプレーンテキスト解析（ingest_paste.parse_records）の検証。"""
    import importlib.util
    spec = importlib.util.spec_from_file_location(
        "ingest_paste",
        Path(__file__).resolve().parent.parent / "scripts" / "ingest_paste.py")
    ip = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(ip)

    print("[paste] 論文プレーンテキストの解析")
    paper_txt = (
        "Circular RNA as a potential biomarker for obstructive sleep apnea.\n"
        "Shinichi Miyazaki, Yu Hayashi\n"
        "Sleep and biological rhythms 23(3) 257-258 2025年7月\n"
    )
    pr = ip.parse_records(paper_txt, "paper")
    check("論文1件を抽出", len(pr) == 1, f"n={len(pr)}")
    r = pr[0]
    check("著者", r.get("authors") == "Shinichi Miyazaki, Yu Hayashi", str(r))
    check("タイトル(末尾ピリオド除去)", r.get("title") == "Circular RNA as a potential biomarker for obstructive sleep apnea")
    check("誌名", r.get("journal") == "Sleep and biological rhythms", str(r.get("journal")))
    check("巻号ページ", r.get("volume") == "23" and r.get("issue") == "3" and r.get("pages") == "257-258")
    check("日付", r.get("date") == "2025/7")

    print("[paste] 発表プレーンテキストの解析（著者行なしを含む）")
    talk_txt = (
        "Exploring regulatory mechanisms and functions of sleep using Caenorhabditis elegans\n"
        "Shinichi Miyazaki, Taizo Kawano, Yu Hayashi\n"
        "日本分子生物学会  2025年12月3日\n"
        "睡眠制御のメカニズムと睡眠の機能の解明に向けて\n"
        "線虫研究の未来を創る会2024  2024年8月28日\n"
    )
    tr = ip.parse_records(talk_txt, "presentation")
    check("発表2件を抽出", len(tr) == 2, f"n={len(tr)}")
    check("発表1: 演題", tr[0].get("title", "").startswith("Exploring regulatory"))
    check("発表1: 学会名", tr[0].get("conference") == "日本分子生物学会")
    check("発表1: 日付", tr[0].get("date") == "2025/12/3")
    check("発表2: 著者行なし→authors空", not tr[1].get("authors"))
    check("発表2: 演題", tr[1].get("title") == "睡眠制御のメカニズムと睡眠の機能の解明に向けて", str(tr[1]))
    check("発表2: 学会名", tr[1].get("conference") == "線虫研究の未来を創る会2024")

    print("[paste] 英語日付（Dec, 2021 等）と月名の途中一致防止")
    en = ip.parse_records(
        "Lipids and proteins changes detected by CARS\n"
        "Shinichi Miyazaki, Hideaki Kano\n"
        "Material Research Meeting  Dec, 2021\n", "presentation")
    check("英語日付1件", len(en) == 1, f"n={len(en)}")
    check("英語日付→2021/12", en and en[0].get("date") == "2021/12", str(en[:1]))
    check("学会名(英語)", en and en[0].get("conference") == "Material Research Meeting")
    check("Marine等の月名途中一致を誤検出しない", ip._find_date("Marine Biology Society 2021") is None)


def llm_parse_tests() -> None:
    """LLM 構造化抽出（scripts/llm_parse.py）の純関数・フォールバック検証（ネットワーク非依存）。"""
    import importlib.util
    spec = importlib.util.spec_from_file_location(
        "llm_parse",
        Path(__file__).resolve().parent.parent / "scripts" / "llm_parse.py")
    lp = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(lp)

    print("[llm] _normalize_llm_records（許可キー除去・型整形・空件除外）")
    raw = [
        # 許可外キー(foo)は除去、数値→文字列＋strip、date は保持。
        {"title": "  REM sleep in mice  ", "authors": "Yamada T", "volume": 16,
         "foo": "x", "date": "2025/7", "doi": ""},
        # タイトルも著者も無い → 除外。
        {"journal": "Sci Rep"},
        # 著者のみ → 採用。
        {"authors": "Hayashi N"},
        # dict 以外 → 無視。
        "not-a-dict",
    ]
    recs = lp._normalize_llm_records(raw, "paper")
    check("空件・非dictを除外して2件", len(recs) == 2, f"n={len(recs)}")
    check("許可外キー foo を除去", "foo" not in recs[0])
    check("数値 volume を文字列化", recs[0].get("volume") == "16")
    check("タイトルを strip", recs[0].get("title") == "REM sleep in mice")
    check("空文字 doi は脱落", "doi" not in recs[0])
    check("著者のみの件を採用", recs[1].get("authors") == "Hayashi N")

    print("[llm] llm_enabled / トークン未設定時のフォールバック")
    check("トークン有→enabled", lp.llm_enabled("ghp_dummy"))
    check("トークン無→disabled", not lp.llm_enabled(""))
    raised = False
    try:
        lp.parse_records_llm("any text", "paper", token="")
    except lp.LLMParseError:
        raised = True
    check("トークン無で LLMParseError（→従来解析へフォールバック）", raised)


def integration_tests(source) -> None:
    print("[integration] loading workbook")
    df = load_publications(source)
    check("non-empty", not df.empty, f"rows={len(df)}")
    present = set(df["type"])
    check("papers present", "paper" in present)
    counts = df["type"].value_counts().to_dict()
    for rtype, label in RECORD_TYPES.items():
        print(f"    - {label}: {counts.get(rtype, 0)} 件")

    # 品質ゲート: status=未確認 の行は読み込まれない。
    titles = " ".join(df["title"].astype(str))
    check("status gate excludes 未確認", "Unverified" not in titles)

    print("[integration] roster")
    roster = load_roster_sheet(source)
    members = parse_roster(roster)
    check("members parsed", len(members) > 5, f"n={len(members)}")
    names = {mm.last for mm in members}
    check("Miyazaki in roster", "Miyazaki" in names)

    print("[integration] author filter (Hayashi)")
    matcher = AuthorMatcher(members, threshold=85)
    hayashi = next((mm for mm in members if mm.last == "Hayashi" and mm.first == "Naoko"), None)
    if hayashi:
        sub = by_authors(df, [hayashi], matcher)
        check("Hayashi has records", len(sub) > 0, f"n={len(sub)}")

    print("[integration] fiscal year filter")
    fys = sorted({int(y) for y in df["fiscal_year"].dropna().unique()})
    if fys:
        sub = by_fiscal_year(df, fys[-1], fys[-1])
        check("latest FY subset <= all", len(sub) <= len(df) and len(sub) >= 0)

    print("[integration] formatting (科研費, papers) + bold/italic + lang")
    templates = load_templates()
    papers = df[df["type"] == "paper"]
    spec = templates["paper"]["科研費"]
    out = render_records(papers, spec, bold_fields={"authors"}, italic_fields={"journal_abbr"})
    check("rendered non-empty", out["count"] > 0 and len(out["plain"]) > 0)
    check("no float .0 in volume", ".0;" not in out["plain"] and ".0)" not in out["plain"])
    check("bold applied in markdown", "**" in out["markdown"])
    check("italic applied in markdown", "*" in out["markdown"])
    check("plain text has no bold markers", "**" not in out["plain"])

    grp = templates["paper"].get("業績一覧（年度別・番号付き）")
    if grp:
        en = render_records(papers, grp, lang="en")
        ja = render_records(papers, grp, lang="ja")
        check("EN fiscal-year heading", "FY " in en["markdown"])
        check("JA fiscal-year heading", "年度" in ja["markdown"])

    print("\n  --- sample (first paper, plain) ---")
    print("  " + (out["plain"].splitlines()[0] if out["plain"] else "(none)"))


_FIXTURE = Path(__file__).resolve().parent / "fixtures" / "canonical_sample.xlsx"


def main() -> None:
    unit_tests()
    author_style_tests()
    form_field_tests(check)
    template_header_tests(check)
    v2_tests()
    paste_tests()
    llm_parse_tests()
    if len(sys.argv) > 1:
        arg = sys.argv[1]
        # ローカルファイルならオフライン検証、それ以外は共有 URL/ID としてライブ取得。
        if Path(arg).exists():
            source = Path(arg).read_bytes()
            print(f"\n[integration] using local file: {arg}")
        else:
            print(f"\n[integration] fetching live: {arg}")
            source = load_workbook_bytes(arg)
    else:
        print(f"\n[integration] using bundled fixture: {_FIXTURE.name}")
        source = _FIXTURE.read_bytes()
    integration_tests(source)

    print(f"\nRESULT: {PASS} passed, {FAIL} failed")
    sys.exit(1 if FAIL else 0)


if __name__ == "__main__":
    main()
