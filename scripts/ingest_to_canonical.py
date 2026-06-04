"""一括取り込み: 既存の業績リストを Canonical（名前付き列）xlsx へ正規化する。

Canonical xlsx は本アプリの新しい入力形式。種別ごとに 1 タブを持ち、
1 行目＝安定したヘッダ名（論理フィールド名）、先頭にメタ列
（record_id, status, submitter, source, created_at）を備える。
Example 行・No 列・列ズレを持たないため、loader.py は列名で素直に読める。

使い方:
    # 1) 初期移行: 現行の崩れた DB（Google Sheet ID / URL / ローカル xlsx）を丸ごと投入
    python scripts/ingest_to_canonical.py --from legacy --out canonical.xlsx
    python scripts/ingest_to_canonical.py --from legacy --src legacy_live.xlsx --out canonical.xlsx

    # 2) 一括入力: メンバーが持つ構造化リスト（1 種別ぶん）を取り込み、既存 Canonical へ追記
    python scripts/ingest_to_canonical.py --from xlsx --type paper --src papers.xlsx --append canonical.xlsx
    python scripts/ingest_to_canonical.py --from csv  --type presentation --src talks.csv --append canonical.xlsx

モード（--from）:
    legacy : 旧フォーマット（位置インデックス）を本スクリプト内蔵の表で抽出。status=確認済。
    xlsx   : ヘッダ名を論理フィールドへ近似マッピング（rapidfuzz）。status=未確認。
    csv    : xlsx と同様（CSV 入力）。

注意:
    本スクリプトは旧フォーマットの列インデックスを **自己完結で内蔵** する。
    パッケージ本体（schema.py/loader.py）は名前ベースへ移行済みのため、ここでしか旧形式を扱わない。
"""

from __future__ import annotations

import argparse
import io
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from rapidfuzz import fuzz

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from publication_summarizer.loader import (  # noqa: E402
    _has_cjk,
    _parse_date,
    extract_sheet_id,
    load_workbook_bytes,
)
from publication_summarizer.schema import BILINGUAL_FIELDS, expand_bilingual  # noqa: E402

# ── Canonical の定義 ───────────────────────────────────────────────
# note: 重複注記（dup_of=...）・補完元（crossref）等のキュレーション用メモ。
META_FIELDS = ["record_id", "status", "submitter", "source", "created_at", "note"]

# 種別ごとの本体列（base 並び）。二ヶ国語 base は expand_bilingual で _ja/_en へ展開。
# schema.SHEET_SPECS と一致（tests/test_form_fields.py が検証）。
_BASE_FIELDS: dict[str, tuple[str, ...]] = {
    "paper": ("date", "category", "peer_reviewed", "authors", "title",
              "journal", "journal_abbr", "volume", "issue", "pages", "doi"),
    "book": ("date", "international", "peer_reviewed", "authors", "review_title",
             "book_title", "chapter", "editor", "volume", "issue", "pages",
             "publisher", "doi", "issn", "isbn"),
    "presentation": ("date", "scope", "title", "authors", "conference",
                     "symposium", "invited", "venue", "presentation_type"),
    "award": ("date", "scope", "authors", "title", "awarded_study", "organization"),
    "outreach": ("date", "scope", "authors", "title", "venue"),
    "publicity": ("date", "media_type", "media_name", "authors", "title", "link"),
}
CANONICAL_FIELDS: dict[str, list[str]] = {
    rt: list(expand_bilingual(bases)) for rt, bases in _BASE_FIELDS.items()
}

# Canonical タブ名（spec_for_sheet のキーワードを含む安定名）。
TAB_NAME: dict[str, str] = {
    "paper": "Original Papers",
    "book": "Books",
    "presentation": "presentations",
    "award": "Awards",
    "outreach": "Outreach",
    "publicity": "Publicity",
}

ID_PREFIX = {"paper": "PAP", "book": "BK", "presentation": "PRE",
             "award": "AWD", "outreach": "OUT", "publicity": "PUB"}

ROSTER_TAB = "Input confirmation"


# ── 旧フォーマット（位置インデックス）の内蔵定義 ─────────────────────
# (keyword, rtype, cols{論理名->旧列index}, title_field)
LEGACY_SPECS = [
    ("Original Papers", "paper", {
        "date": 2, "category": 3, "peer_reviewed": 4, "authors": 5, "title": 6,
        "journal": 7, "journal_abbr": 8, "volume": 9, "issue": 10, "pages": 11,
        "doi": 12}, "title"),
    ("Books", "book", {
        "date": 2, "international": 3, "peer_reviewed": 4, "authors": 5,
        "review_title": 6, "book_title": 7, "chapter": 8, "editor": 9,
        "volume": 10, "issue": 11, "pages": 12, "publisher": 13, "doi": 14,
        "issn": 15, "isbn": 16}, "book_title"),
    ("presentations", "presentation", {
        "date": 2, "scope": 3, "title": 4, "authors": 5, "conference": 6,
        "symposium": 7, "invited": 8, "venue": 9, "presentation_type": 12},
        "title"),
    ("Awards", "award", {
        "date": 2, "scope": 3, "authors": 4, "title": 5, "awarded_study": 6,
        "organization": 7}, "title"),
    ("Outreach", "outreach", {
        "date": 2, "scope": 3, "authors": 4, "title": 5, "venue": 6}, "title"),
    ("Publicity", "publicity", {
        "date": 2, "media_type": 3, "media_name": 4, "authors": 5, "title": 6,
        "link": 7}, "title"),
]

_NO_IDX = 0       # A列 = No
_MARKER_IDX = 1   # B列 = Example マーカー（"例"）


def _legacy_spec(sheet_name: str):
    low = sheet_name.lower()
    for keyword, rtype, cols, title_field in LEGACY_SPECS:
        if keyword.lower() in low:
            return rtype, cols, title_field
    return None


def _raw_cell(row: pd.Series, idx: int):
    """旧シートのセルを取り出す（datetime/数値は保持、NaN は ''）。"""
    if idx >= len(row):
        return ""
    val = row.iloc[idx]
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return ""
    if isinstance(val, str):
        return val.strip()
    return val


def _is_real_row(no_value, marker_value) -> bool:
    """実データ行か判定（Example 行・空行・見出し行を除外）。loader 旧ロジック準拠。"""
    import re
    if isinstance(marker_value, str) and marker_value.strip() == "例":
        return False
    if no_value is None or (isinstance(no_value, float) and pd.isna(no_value)):
        return False
    if isinstance(no_value, str):
        no_value = no_value.strip()
        if not no_value or no_value.lower() == "example":
            return False
        return bool(re.match(r"^\d+(\.\d+)?$", no_value))
    return isinstance(no_value, (int, float))


# ── 取り込みロジック ─────────────────────────────────────────────
def ingest_legacy(src_bytes: bytes) -> tuple[dict[str, list[dict]], pd.DataFrame | None]:
    """旧フォーマット xlsx から種別別レコードと名簿シートを抽出。"""
    sheets = pd.read_excel(io.BytesIO(src_bytes), sheet_name=None, header=None)
    by_type: dict[str, list[dict]] = {rt: [] for rt in CANONICAL_FIELDS}
    roster_raw: pd.DataFrame | None = None

    for name, raw in sheets.items():
        if ROSTER_TAB.lower() in name.lower():
            roster_raw = raw
            continue
        spec = _legacy_spec(name)
        if spec is None:
            continue
        rtype, cols, title_field = spec
        for _, row in raw.iterrows():
            no_val = row.iloc[_NO_IDX] if _NO_IDX < len(row) else None
            marker = row.iloc[_MARKER_IDX] if _MARKER_IDX < len(row) else None
            if not _is_real_row(no_val, marker):
                continue
            rec = {f: _raw_cell(row, cols[f]) for f in CANONICAL_FIELDS[rtype] if f in cols}
            authors = str(rec.get("authors", "")).strip()
            title = str(rec.get(title_field, "")).strip()
            date_cell = str(rec.get("date", "")).strip()
            if not (authors or title or date_cell):
                continue
            by_type[rtype].append(rec)
    return by_type, roster_raw


# ヘッダ近似マッピング用の別名（論理フィールド -> 受理する表記の候補）。
FIELD_ALIASES: dict[str, list[str]] = {
    "date": ["発行日", "発表日", "受賞日", "実施日", "掲載日", "年月日", "date", "year", "発行年"],
    "category": ["区分", "category"],
    "peer_reviewed": ["査読", "査読の有無", "peer reviewed", "peer-reviewed"],
    "international": ["国内国際", "国内/国際", "international"],
    "scope": ["国内国際", "国内/国際", "scope", "domestic international"],
    "authors": ["著者", "発表者", "受賞者", "実施者", "掲載人物", "authors", "author"],
    "title": ["タイトル", "演題", "賞の名称", "活動概要", "掲載概要", "title", "論文タイトル"],
    "journal": ["雑誌名", "雑誌名正式", "journal", "雑誌"],
    "journal_abbr": ["略誌名", "雑誌名略", "journal abbr", "abbreviation"],
    "volume": ["巻", "volume", "vol"],
    "issue": ["号", "issue", "no"],
    "pages": ["ページ", "頁", "pages", "page"],
    "doi": ["doi"],
    "review_title": ["章総説タイトル", "総説タイトル", "review title"],
    "book_title": ["書名", "書籍名", "book title"],
    "chapter": ["章", "chapter"],
    "editor": ["編者", "editor"],
    "publisher": ["出版社", "publisher"],
    "issn": ["issn"],
    "isbn": ["isbn"],
    "conference": ["学会名", "学会研究会名", "conference", "学会"],
    "symposium": ["シンポジウム", "シンポジウム名", "symposium"],
    "invited": ["招待", "招待の有無", "invited"],
    "venue": ["開催地", "会場", "venue", "開催地媒体"],
    "presentation_type": ["発表形式", "presentation type", "口頭ポスター"],
    "awarded_study": ["受賞対象研究", "受賞対象", "awarded study"],
    "organization": ["授与団体", "団体", "organization"],
    "media_type": ["媒体種別", "media type"],
    "media_name": ["媒体名", "media name"],
    "link": ["リンク", "url", "link"],
}


def _norm_header(text: str) -> str:
    import re
    import unicodedata
    text = unicodedata.normalize("NFKC", str(text))
    text = re.sub(r"[\s_/\-（）()・,.:;]", "", text)
    return text.strip().lower()


# 言語マーカー（見出し内に含まれていれば二ヶ国語 base の片側へ振り分ける）。
_JA_MARKERS = ["日本語", "和文", "邦文", "japanese", "ja", "jp"]
_EN_MARKERS = ["英語", "英文", "english", "en"]


def _detect_marker(norm: str) -> tuple[str | None, str]:
    """正規化ヘッダから言語マーカーを検出し、(lang, マーカー除去後) を返す。"""
    for mk in _JA_MARKERS:
        if mk in norm:
            return "ja", norm.replace(mk, "")
    for mk in _EN_MARKERS:
        if mk in norm:
            return "en", norm.replace(mk, "")
    return None, norm


def _match_field(header: str, rtype: str, threshold: int = 80) -> str | None:
    """入力ヘッダを、その種別の Canonical 列へ近似マッピング。

    base 名で照合し、二ヶ国語 base は見出しの言語マーカー（日本語/英語 等）で
    `_ja`/`_en` を決める。マーカーが無い汎用見出しは base を返し、後段
    （_ensure_bilingual）が本文の言語で振り分ける。
    """
    norm = _norm_header(header)
    if not norm:
        return None
    lang, stripped = _detect_marker(norm)
    best_base, best_score = None, 0
    for base in _BASE_FIELDS[rtype]:
        candidates = [base] + FIELD_ALIASES.get(base, [])
        cand_norms = [_norm_header(c) for c in candidates]
        if stripped in cand_norms or norm in cand_norms:
            best_base = base
            break
        score = max(fuzz.ratio(stripped, c) for c in cand_norms)
        if score > best_score:
            best_base, best_score = base, score
    else:
        if best_score < threshold:
            return None
    if best_base is None:
        return None
    if best_base in BILINGUAL_FIELDS:
        if lang == "ja":
            return best_base + "_ja"
        if lang == "en":
            return best_base + "_en"
        return best_base  # 汎用見出し → 後段で言語振り分け
    return best_base


def ingest_structured(src: Path, rtype: str, is_csv: bool) -> list[dict]:
    """構造化リスト（xlsx/csv・1 種別）を読み、論理フィールドへマッピング。"""
    df = pd.read_csv(src) if is_csv else pd.read_excel(src, header=0)
    colmap: dict[str, str] = {}
    for col in df.columns:
        field = _match_field(str(col), rtype)
        if field and field not in colmap.values():
            colmap[col] = field
    matched = sorted(set(colmap.values()))
    unmatched = [str(c) for c in df.columns if c not in colmap]
    print(f"  マッピング成立: {matched}")
    if unmatched:
        print(f"  [警告] 未対応の列（無視）: {unmatched}")
    if "authors" not in matched and "title" not in matched:
        print("  [警告] authors/title いずれもマッピングできませんでした。ヘッダ名を確認してください。")

    records: list[dict] = []
    for _, row in df.iterrows():
        rec = {field: ("" if pd.isna(row[col]) else row[col]) for col, field in colmap.items()}
        if not (str(rec.get("authors", "")).strip() or _rec_title(rec) or
                str(rec.get("title", "")).strip()):
            continue
        records.append(rec)
    return records


# ── 出力 ─────────────────────────────────────────────────────────
def _cell_out(value):
    """xlsx セルへ書ける形へ整える（Timestamp は ISO 文字列）。"""
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    if isinstance(value, (pd.Timestamp, datetime)):
        ts = _parse_date(value)
        return ts.strftime("%Y/%m/%d") if not pd.isna(ts) else ""
    return value


def _ensure_bilingual(rec: dict) -> dict:
    """汎用 base 値（title 等）を本文の言語で `_ja`/`_en` へ振り分ける。

    既に `_ja`/`_en` が入っていればそれを優先（上書きしない）。rec は破壊しない。
    """
    out = dict(rec)
    for base in BILINGUAL_FIELDS:
        if base not in out:
            continue
        val = out.pop(base)
        val_s = str(val).strip() if not (isinstance(val, float) and pd.isna(val)) else ""
        if not val_s:
            continue
        slot = base + ("_ja" if _has_cjk(val_s) else "_en")
        if not str(out.get(slot, "")).strip():
            out[slot] = val
    return out


def _s(value) -> str:
    """セル値を安全に文字列化（NaN/None は空文字。str(NaN)=='nan' の混入を防ぐ）。"""
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    return str(value).strip()


def _rec_title(rec: dict) -> str:
    """重複判定・存在チェック用の代表タイトル（_ja/_en/base/book_title を横断）。"""
    for key in ("title_ja", "title_en", "title",
                "book_title_ja", "book_title_en", "book_title"):
        v = _s(rec.get(key))
        if v:
            return v
    return ""


def _ym(value) -> str:
    """日付を年/月（YYYY/MM）に正規化。形式差（2025/7 と 2025/07/01）を吸収。"""
    ts = _parse_date(value)
    if not (ts is None or pd.isna(ts)):
        return ts.strftime("%Y/%m")
    return str(_cell_out(value)).strip()


def _dup_keys(rec: dict) -> tuple[str | None, str]:
    """重複検出キーの対 (doi_key, dt_key) を返す。

    doi が一致、または「年月＋タイトル先頭」が一致すれば重複とみなす（どちらか一方でも可）。
    これにより DOI 有無の混在（フォーム⇄貼り付け）や日付表記ゆれでも検出できる。
    """
    doi = _s(rec.get("doi")).lower()
    doi_key = f"doi:{doi}" if doi else None
    title = _rec_title(rec).lower()[:40]
    dt_key = f"dt:{_ym(rec.get('date', ''))}|{title}"
    return doi_key, dt_key


def _read_existing(path: Path) -> dict[str, list[dict]]:
    """既存 Canonical を種別別レコードへ読み戻す（--append 用）。"""
    by_type: dict[str, list[dict]] = {rt: [] for rt in CANONICAL_FIELDS}
    if not path.exists():
        return by_type
    sheets = pd.read_excel(path, sheet_name=None, header=0)
    name_to_type = {v.lower(): k for k, v in TAB_NAME.items()}
    for name, df in sheets.items():
        rtype = name_to_type.get(name.lower())
        if rtype is None:
            continue
        for _, row in df.iterrows():
            by_type[rtype].append({c: row[c] for c in df.columns})
    return by_type


def _next_seq(existing: list[dict]) -> int:
    return len(existing) + 1


def write_canonical(out: Path, by_type: dict[str, list[dict]], source_label: str,
                    status: str, roster_raw: pd.DataFrame | None = None,
                    append_to: Path | None = None) -> dict[str, int]:
    """Canonical xlsx を書き出す（append_to 指定時は既存へマージ・重複除外）。"""
    base = _read_existing(append_to) if append_to else {rt: [] for rt in CANONICAL_FIELDS}
    added: dict[str, int] = {}
    now = datetime.now().strftime("%Y/%m/%d %H:%M:%S")

    for rtype, new_recs in by_type.items():
        existing = base[rtype]
        seen_doi: set[str] = set()
        seen_dt: set[str] = set()
        for r in existing:
            dk, tk = _dup_keys(r)
            if dk:
                seen_doi.add(dk)
            seen_dt.add(tk)
        prefix = ID_PREFIX[rtype]
        seq = _next_seq(existing)
        n_added = 0
        for rec in new_recs:
            rec = _ensure_bilingual(rec)
            doi_key, dt_key = _dup_keys(rec)
            if (doi_key and doi_key in seen_doi) or dt_key in seen_dt:
                continue
            if doi_key:
                seen_doi.add(doi_key)
            seen_dt.add(dt_key)
            full = {f: _cell_out(rec.get(f, "")) for f in CANONICAL_FIELDS[rtype]}
            full.update({
                "record_id": f"{prefix}-{seq:04d}",
                "status": status,
                "submitter": str(rec.get("submitter", "")).strip(),
                "source": source_label,
                "created_at": now,
                "note": str(rec.get("note", "")).strip(),
            })
            existing.append(full)
            seq += 1
            n_added += 1
        added[rtype] = n_added

    _save_workbook(out, base, roster_raw)
    return added


def _save_workbook(out: Path, by_type: dict[str, list[dict]],
                   roster_raw: pd.DataFrame | None = None) -> None:
    """種別別レコード（full 行 dict）を v2 Canonical xlsx として保存する。"""
    wb = Workbook()
    wb.remove(wb.active)
    for rtype, fields in CANONICAL_FIELDS.items():
        ws = wb.create_sheet(TAB_NAME[rtype])
        header = META_FIELDS + fields
        ws.append(header)
        for rec in by_type[rtype]:
            ws.append([rec.get(h, "") for h in header])

    # 名簿タブ（あれば複写）。parse_roster は B列=役職, C列=氏名 を読む。
    if roster_raw is not None and not roster_raw.empty:
        ws = wb.create_sheet(ROSTER_TAB)
        for _, row in roster_raw.iterrows():
            ws.append([("" if pd.isna(v) else v) for v in row.tolist()])

    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)


def upgrade_canonical(src: Path, out: Path) -> dict[str, int]:
    """旧 Canonical（単一言語列）を v2（_ja/_en + note）へ移行する。

    メタ（record_id/status/submitter/source/created_at）は保持し、二ヶ国語 base は
    本文の言語で _ja/_en へ振り分ける。再採番・status 変更・重複除外はしない。
    """
    sheets = pd.read_excel(src, sheet_name=None, header=0)
    raw_sheets = pd.read_excel(src, sheet_name=None, header=None)
    name_to_type = {v.lower(): k for k, v in TAB_NAME.items()}
    by_type: dict[str, list[dict]] = {rt: [] for rt in CANONICAL_FIELDS}
    counts: dict[str, int] = {}
    now = datetime.now().strftime("%Y/%m/%d %H:%M:%S")

    for name, df in sheets.items():
        rtype = name_to_type.get(name.lower())
        if rtype is None:
            continue
        seq = 1
        for _, row in df.iterrows():
            rec = {c: ("" if pd.isna(row[c]) else row[c]) for c in df.columns}
            rec = _ensure_bilingual(rec)
            if not (str(rec.get("authors", "")).strip() or _rec_title(rec)
                    or str(_cell_out(rec.get("date", ""))).strip()):
                continue
            rid = str(rec.get("record_id", "")).strip() or f"{ID_PREFIX[rtype]}-{seq:04d}"
            full = {f: _cell_out(rec.get(f, "")) for f in CANONICAL_FIELDS[rtype]}
            full.update({
                "record_id": rid,
                "status": str(rec.get("status", "")).strip(),
                "submitter": str(rec.get("submitter", "")).strip(),
                "source": str(rec.get("source", "")).strip() or "upgrade",
                "created_at": str(rec.get("created_at", "")).strip() or now,
                "note": str(rec.get("note", "")).strip(),
            })
            by_type[rtype].append(full)
            seq += 1
        counts[rtype] = len(by_type[rtype])

    roster_raw = next((raw for nm, raw in raw_sheets.items()
                       if ROSTER_TAB.lower() in nm.lower()), None)
    _save_workbook(out, by_type, roster_raw)
    return counts


def main() -> None:
    ap = argparse.ArgumentParser(description="既存業績リスト → Canonical xlsx 取り込み")
    ap.add_argument("--from", dest="mode", required=True, choices=["legacy", "xlsx", "csv", "upgrade"])
    ap.add_argument("--src", help="入力ソース: legacy=Sheet ID/URL もしくはローカル xlsx / xlsx,csv,upgrade=ファイルパス")
    ap.add_argument("--type", dest="rtype", choices=list(CANONICAL_FIELDS),
                    help="xlsx/csv モードで必須: 取り込む業績種別")
    ap.add_argument("--out", help="出力 Canonical xlsx（新規作成）")
    ap.add_argument("--append", help="既存 Canonical xlsx に追記（重複除外）。--out 省略時はこのファイルへ上書き")
    args = ap.parse_args()

    append_to = Path(args.append) if args.append else None
    out = Path(args.out) if args.out else append_to
    if out is None:
        ap.error("--out か --append のいずれかが必要です。")

    if args.mode == "upgrade":
        if not args.src or not Path(args.src).exists():
            ap.error("upgrade モードでは既存 Canonical xlsx を --src で指定してください。")
        counts = upgrade_canonical(Path(args.src), out)
        total = sum(counts.values())
        print(f"\n[完了] {out} へ v2（_ja/_en + note）として {total} 件を移行しました。")
        for rtype, n in counts.items():
            if n:
                print(f"  - {TAB_NAME[rtype]}: {n}")
        return

    if args.mode == "legacy":
        src = args.src
        if src and Path(src).exists():
            src_bytes = Path(src).read_bytes()
            print(f"[legacy] ローカル読み込み: {src}")
        else:
            from publication_summarizer.loader import DEFAULT_SHEET_ID
            sid = extract_sheet_id(src) if src else DEFAULT_SHEET_ID
            print(f"[legacy] ライブ取得: {sid}")
            src_bytes = load_workbook_bytes(sid)
        by_type, roster = ingest_legacy(src_bytes)
        added = write_canonical(out, by_type, source_label="legacy", status="確認済",
                                roster_raw=roster, append_to=append_to)
    else:
        if not args.rtype:
            ap.error("xlsx/csv モードでは --type が必須です。")
        if not args.src:
            ap.error("xlsx/csv モードでは --src が必須です。")
        recs = ingest_structured(Path(args.src), args.rtype, is_csv=(args.mode == "csv"))
        by_type = {rt: [] for rt in CANONICAL_FIELDS}
        by_type[args.rtype] = recs
        added = write_canonical(out, by_type, source_label=args.mode, status="未確認",
                                append_to=append_to)

    total = sum(added.values())
    print(f"\n[完了] {out} へ {total} 件を取り込みました（status={'確認済' if args.mode=='legacy' else '未確認'}）")
    for rtype, n in added.items():
        if n:
            print(f"  - {TAB_NAME[rtype]}: +{n}")


if __name__ == "__main__":
    main()
