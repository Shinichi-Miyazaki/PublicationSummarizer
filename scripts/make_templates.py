"""一括入力テンプレート（Excel）を生成する。

種別ごとに「正しい見出し」を持つ記入用シートと、「記入例・注意」シートを備えた
xlsx を作る。メンバーはこの表に多件まとめて記入して返し、管理者は既存の
ingest_to_canonical.py で本 DB（Canonical）へ一括追記する:

    # 例: 受領した papers_filled.xlsx を取り込み（重複除外・status=未確認）
    python scripts/ingest_to_canonical.py --from xlsx --type paper \\
        --src papers_filled.xlsx --append canonical.xlsx

設計:
    - 1 枚目のシート「入力」… 見出しのみ（メンバーは 2 行目以降に記入）。
      ingest_to_canonical は xlsx の **1 枚目だけ** を読むため、記入例が混入しない。
    - 2 枚目のシート「記入例・注意」… 必須項目・日付書式・選択肢・例を表示（取込対象外）。
    - 見出し語は ingest の FIELD_ALIASES と完全一致する文言を選び、確実にマッピングさせる
      （tests/test_form_fields.py が見出し→論理フィールドの対応を自動検証）。
    - 必須/選択肢は scripts/forms/publication_form.gs の FIELD_MAP を出典とし、フォームと統一。

使い方:
    python scripts/make_templates.py --all                 # 全 6 種別を templates/ に生成
    python scripts/make_templates.py --type paper          # 指定種別のみ
    python scripts/make_templates.py --all --outdir out    # 出力先を変更
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

sys.path.insert(0, str(Path(__file__).resolve().parent))  # ingest_to_canonical を読む

from ingest_to_canonical import CANONICAL_FIELDS, TAB_NAME  # noqa: E402
from publication_summarizer.schema import BILINGUAL_FIELDS  # noqa: E402

_GS_PATH = Path(__file__).resolve().parent / "forms" / "publication_form.gs"

# base 見出し語（base → メンバー向け日本語見出し）。各見出しは ingest の FIELD_ALIASES と
# 一致する文言。二ヶ国語 base は下の _expand で「○○（日本語）」「○○（英語）」の2列へ展開する。
BASE_LABELS: dict[str, dict[str, str]] = {
    "paper": {
        "date": "発行日", "category": "区分", "peer_reviewed": "査読の有無",
        "authors": "著者", "title": "論文タイトル", "journal": "雑誌名",
        "journal_abbr": "略誌名", "volume": "巻", "issue": "号", "pages": "ページ",
        "doi": "DOI",
    },
    "book": {
        "date": "発行日", "international": "国内/国際", "peer_reviewed": "査読の有無",
        "authors": "著者", "review_title": "章・総説タイトル", "book_title": "書名",
        "chapter": "章", "editor": "編者", "volume": "巻", "issue": "号",
        "pages": "ページ", "publisher": "出版社", "doi": "DOI", "issn": "ISSN",
        "isbn": "ISBN",
    },
    "presentation": {
        "date": "発表日", "scope": "国内/国際", "title": "演題", "authors": "発表者",
        "conference": "学会・研究会名", "symposium": "シンポジウム名",
        "invited": "招待の有無", "venue": "開催地", "presentation_type": "発表形式",
    },
    "award": {
        "date": "受賞日", "scope": "国内/国際", "authors": "受賞者", "title": "賞の名称",
        "awarded_study": "受賞対象", "organization": "授与団体",
    },
    "outreach": {
        "date": "実施日", "scope": "国内/国際", "authors": "実施者",
        "title": "活動概要", "venue": "開催地・媒体",
    },
    "publicity": {
        "date": "掲載日", "media_type": "媒体種別", "media_name": "媒体名",
        "authors": "掲載人物", "title": "掲載概要", "link": "リンク",
    },
}

# base 記入例（二ヶ国語 base は (日本語例, 英語例) のタプル）。注意シートにのみ表示。
BASE_EXAMPLES: dict[str, dict[str, object]] = {
    "paper": {
        "date": "2026/04/10", "category": "原著論文", "peer_reviewed": "査読あり",
        "authors": "山田 太郎, 林 直子, 鈴木 一郎",
        "title": ("マウスの社会的順位とレム睡眠", "Social rank affects REM sleep in mice"),
        "journal": ("", "Scientific Reports"), "journal_abbr": ("", "Sci. Rep."),
        "volume": "16", "issue": "1", "pages": "871", "doi": "10.1038/s41598-025-32402-2",
    },
    "book": {
        "date": "2025/11/01", "international": "国内", "peer_reviewed": "査読なし",
        "authors": "山田 太郎",
        "review_title": ("睡眠覚醒の制御機構", ""), "book_title": ("最新・睡眠科学", ""),
        "chapter": "3", "editor": "鈴木 一郎", "volume": "", "issue": "", "pages": "45-60",
        "publisher": "医学書院", "doi": "", "issn": "", "isbn": "978-4-260-00000-0",
    },
    "presentation": {
        "date": "2025/09/20", "scope": "国内",
        "title": ("睡眠とストレスの関係", "Sleep and stress"),
        "authors": "山田 太郎, 林 直子",
        "conference": ("日本神経科学大会", "Annual Meeting of JNS"),
        "symposium": ("", ""), "invited": "招待なし", "venue": "横浜", "presentation_type": "口頭",
    },
    "award": {
        "date": "2025/12/05", "scope": "国内", "authors": "山田 太郎",
        "title": ("若手奨励賞", "Young Investigator Award"),
        "awarded_study": "睡眠制御に関する一連の研究", "organization": "日本睡眠学会",
    },
    "outreach": {
        "date": "2025/08/03", "scope": "国内", "authors": "山田 太郎",
        "title": ("市民講座「睡眠のふしぎ」", ""), "venue": "市民会館",
    },
    "publicity": {
        "date": "2026/01/15", "media_type": "新聞", "media_name": "○○新聞",
        "authors": "山田 太郎", "title": ("研究室の睡眠研究が紹介された", ""),
        "link": "https://example.com/article",
    },
}


def _expand_labels(base_map: dict[str, str]) -> dict[str, str]:
    """base ラベルを v2 列名キーへ展開（二ヶ国語は （日本語）/（英語） を付す）。"""
    out: dict[str, str] = {}
    for base, label in base_map.items():
        if base in BILINGUAL_FIELDS:
            out[base + "_ja"] = f"{label}（日本語）"
            out[base + "_en"] = f"{label}（英語）"
        else:
            out[base] = label
    return out


def _expand_examples(base_map: dict[str, object]) -> dict[str, str]:
    """base 記入例を v2 列名キーへ展開（タプルは ja/en へ割当）。"""
    out: dict[str, str] = {}
    for base, ex in base_map.items():
        if base in BILINGUAL_FIELDS:
            ja, en = ex if isinstance(ex, tuple) else (ex, "")
            out[base + "_ja"] = ja
            out[base + "_en"] = en
        else:
            out[base] = ex
    return out


HEADER_LABELS: dict[str, dict[str, str]] = {rt: _expand_labels(m) for rt, m in BASE_LABELS.items()}
EXAMPLES: dict[str, dict[str, str]] = {rt: _expand_examples(m) for rt, m in BASE_EXAMPLES.items()}

_HEADER_FILL = PatternFill("solid", fgColor="DDEBF7")
_NOTE_FILL = PatternFill("solid", fgColor="FFF2CC")


def _load_field_map() -> dict:
    """publication_form.gs の FIELD_MAP（必須・選択肢の出典）を読む。"""
    text = _GS_PATH.read_text(encoding="utf-8")
    m = re.search(r"// FIELD_MAP_JSON_BEGIN\s*var FIELD_MAP\s*=\s*(\{.*?\})\s*;\s*// FIELD_MAP_JSON_END",
                  text, re.DOTALL)
    if not m:
        raise ValueError("publication_form.gs から FIELD_MAP を抽出できませんでした。")
    return json.loads(m.group(1))


def build_template(rtype: str, field_map: dict, outdir: Path) -> Path:
    """1 種別のテンプレート xlsx を作って保存パスを返す。"""
    fields = CANONICAL_FIELDS[rtype]
    labels = HEADER_LABELS[rtype]
    spec = field_map[rtype]
    q_by_field = {q["field"]: q for q in spec["questions"]}

    wb = Workbook()

    # ── 1 枚目: 入力（見出しのみ） ──
    ws = wb.active
    ws.title = "入力"
    headers = [labels[f] for f in fields]
    ws.append(headers)
    for col, f in enumerate(fields, start=1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.fill = _HEADER_FILL
        ws.column_dimensions[cell.column_letter].width = max(10, len(labels[f]) * 2 + 4)

    # ── 2 枚目: 記入例・注意（取込対象外） ──
    note = wb.create_sheet("記入例・注意")
    note.append([f"【{spec['label']}】一括入力テンプレート"])
    note["A1"].font = Font(bold=True, size=12)
    note.append(["・1 枚目「入力」シートの 2 行目以降に、1 行＝1 件で記入してください。"])
    note.append(["・見出し（1 行目）は変更しないでください。列の追加・削除も不要です。"])
    note.append(["・日付は YYYY/MM/DD 形式（例 2026/04/10）。"])
    required = [labels[q["field"]] for q in spec["questions"] if q.get("required")]
    note.append([f"・必須項目: {', '.join(required)}"])
    choice_lines = [f"・「{labels[q['field']]}」の選択肢: {' / '.join(q['choices'])}"
                    for q in spec["questions"] if q.get("choices")]
    for line in choice_lines:
        note.append([line])
    for r in range(2, note.max_row + 1):
        note.cell(row=r, column=1).fill = _NOTE_FILL

    note.append([])
    note.append(["▼ 記入例（この行は提出ファイルには不要・参考用）"])
    note.cell(row=note.max_row, column=1).font = Font(bold=True)
    note.append(headers)
    for cell in note[note.max_row]:
        cell.font = Font(bold=True)
    note.append([EXAMPLES[rtype].get(f, "") for f in fields])
    note.column_dimensions["A"].width = 60

    outdir.mkdir(parents=True, exist_ok=True)
    path = outdir / f"{rtype}_template.xlsx"
    wb.save(path)
    return path


def main() -> None:
    ap = argparse.ArgumentParser(description="一括入力テンプレート（Excel）を生成")
    ap.add_argument("--type", dest="rtype", choices=list(CANONICAL_FIELDS),
                    help="生成する種別（省略時は --all）")
    ap.add_argument("--all", action="store_true", help="全 6 種別を生成")
    ap.add_argument("--outdir", default="templates", help="出力先フォルダ（既定: templates）")
    args = ap.parse_args()

    if not args.rtype and not args.all:
        ap.error("--type か --all のいずれかを指定してください。")

    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:  # noqa: BLE001
        pass

    field_map = _load_field_map()
    outdir = Path(args.outdir)
    types = list(CANONICAL_FIELDS) if args.all else [args.rtype]

    for rtype in types:
        path = build_template(rtype, field_map, outdir)
        print(f"  [生成] {TAB_NAME[rtype]:16s} -> {path}")
    print(f"\n[完了] {len(types)} 種別のテンプレートを {outdir}/ に生成しました。")
    print("  記入後の取込: python scripts/ingest_to_canonical.py --from xlsx "
          "--type <種別> --src <記入済ファイル> --append canonical.xlsx")


if __name__ == "__main__":
    main()
